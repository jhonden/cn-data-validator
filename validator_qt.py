#!/usr/bin/env python3
"""
Network Package Validator - PyQt6 GUI Version
A local tool for validating network data collection results.
Helps users quickly identify issues with invalid file formats and missing files.

Implemented using PyQt6 for better cross-platform compatibility (no tkinter dependency).
"""

import sys
import os
import traceback
from datetime import datetime

# 添加当前目录到sys.path以支持模块导入
if os.path.dirname(os.path.abspath(__file__)) not in sys.path:
    sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from PyQt6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QFileDialog, QTableWidget, QTableWidgetItem,
    QHeaderView, QMessageBox, QProgressBar, QStatusBar, QFrame,
    QLineEdit, QComboBox, QDialog
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal
from PyQt6.QtGui import QColor, QFont
from utils.file_scanner import FileScanner
from utils.package_identifier import PackageIdentifier
from utils.design_tokens import COLORS, SPACING, SIZES, BORDER_RADIUS
from utils.typography import FONT_SIZE, FONT_WEIGHT

# Excel export support
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
from exceptions import (
    ValidationException,
    DirectoryNotFoundException,
    PermissionDeniedException,
    InvalidPackageException,
    FileSystemException,
    MemoryException
)


class ValidationThread(QThread):
    """Validation thread to avoid blocking UI"""
    finished = pyqtSignal(object, str)
    progress = pyqtSignal(int)

    def __init__(self, directory):
        super().__init__()
        self.directory = directory

    def run(self):
        try:
            # Check if directory exists
            if not os.path.exists(self.directory):
                raise DirectoryNotFoundException(self.directory)

            # Check if directory is accessible
            if not os.access(self.directory, os.R_OK):
                raise PermissionDeniedException(self.directory)

            scanner = FileScanner(self.directory)
            # 传递进度回调，发送实际进度信号
            scanner.scan_directory(progress_callback=lambda p: self.progress.emit(p))
            self.finished.emit(scanner, "")

        except (DirectoryNotFoundException, PermissionDeniedException,
                InvalidPackageException, FileSystemException, MemoryException) as e:
            # 自定义异常，直接传递
            self.finished.emit(None, e)

        except PermissionError as e:
            # 权限错误转换为自定义异常
            self.finished.emit(None, PermissionDeniedException(str(e)))

        except FileNotFoundError as e:
            # 文件未找到转换为自定义异常
            self.finished.emit(None, DirectoryNotFoundException(str(e)))

        except MemoryError:
            # 内存不足
            self.finished.emit(None, MemoryException())

        except OSError as e:
            # 其他文件系统错误
            self.finished.emit(None, FileSystemException(str(e)))

        except Exception as e:
            # 其他未知错误
            import traceback as tb
            error_details = f"{str(e)}\n\n{tb.format_exc()}"
            self.finished.emit(None, error_details)


class ValidatorApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.current_dir = ""
        self.validation_thread = None
        self.scanner = None

        # Filter related attributes
        self.all_valid_files = []  # 存储所有有效文件数据
        self.all_invalid_files = []  # 存储所有无效文件数据
        self.filter_widgets = {}  # 存储筛选控件

        self.init_ui()

    def init_ui(self):
        """Initialize UI"""
        self.setWindowTitle("Network Package Validator")
        self.setGeometry(100, 100, 1200, 700)

        # Main window
        main_widget = QWidget()
        self.setCentralWidget(main_widget)

        # Main layout
        layout = QVBoxLayout()
        main_widget.setLayout(layout)

        # Top control area
        control_layout = QHBoxLayout()

        self.dir_label = QLabel("Select Package Directory:")
        self.dir_label.setFont(QFont('Arial', FONT_SIZE['small']))
        control_layout.addWidget(self.dir_label)

        self.dir_path_label = QLabel("")
        self.dir_path_label.setStyleSheet(
            f"color: {COLORS['text_primary']}; "
            f"font-weight: {FONT_WEIGHT['bold']}; "
            f"background-color: {COLORS['info']}; "
            f"padding: {SPACING['small']}px {SPACING['medium']}px; "
            f"border-radius: {BORDER_RADIUS['tiny']}px;"
        )
        self.dir_path_label.setMinimumWidth(SIZES['input_medium'])
        control_layout.addWidget(self.dir_path_label)

        self.select_btn = QPushButton("Select Directory")
        self.select_btn.setFixedWidth(SIZES['button_medium'])
        self.select_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {COLORS['primary']};
                color: {COLORS['text_inverse']};
                border: none;
                padding: {SPACING['small']}px {SPACING['medium']}px;
                border-radius: {BORDER_RADIUS['tiny']}px;
            }}
            QPushButton:hover {{
                background-color: {COLORS['link_hover']};
            }}
        """)
        self.select_btn.clicked.connect(self.select_directory)
        control_layout.addWidget(self.select_btn)

        self.validate_btn = QPushButton("Start Validation")
        self.validate_btn.setFixedWidth(SIZES['button_medium'])
        self.validate_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {COLORS['success']};
                color: {COLORS['text_inverse']};
                border: none;
                padding: {SPACING['small']}px {SPACING['medium']}px;
                border-radius: {BORDER_RADIUS['tiny']}px;
                font-weight: {FONT_WEIGHT['bold']};
            }}
            QPushButton:hover {{
                background-color: #45a049;
            }}
            QPushButton:disabled {{
                background-color: {COLORS['text_hint']};
            }}
        """)
        self.validate_btn.clicked.connect(self.start_validation)
        self.validate_btn.setEnabled(False)
        control_layout.addWidget(self.validate_btn)

        self.progress_bar = QProgressBar()
        self.progress_bar.setFixedWidth(200)
        self.progress_bar.setVisible(False)
        control_layout.addWidget(self.progress_bar)

        control_layout.addStretch()
        layout.addLayout(control_layout)

        # Separator
        line = QFrame()
        line.setFrameShape(QFrame.Shape.HLine)
        line.setFrameShadow(QFrame.Shadow.Sunken)
        layout.addWidget(line)

        # Filter bar
        filter_layout = QHBoxLayout()

        # Filter label
        filter_label = QLabel("Filter:")
        filter_label.setFont(QFont('Arial', FONT_SIZE['small'], QFont.Weight.Bold))
        filter_label.setStyleSheet(f"color: {COLORS['text_primary']};")
        filter_layout.addWidget(filter_label)

        # Filename filter
        self.filter_widgets['filename'] = QLineEdit()
        self.filter_widgets['filename'].setPlaceholderText("Filename...")
        self.filter_widgets['filename'].setFixedWidth(150)
        self.filter_widgets['filename'].textChanged.connect(self.apply_filters)
        filter_layout.addWidget(QLabel("Name:"))
        filter_layout.addWidget(self.filter_widgets['filename'])

        # Format filter (ComboBox)
        self.filter_widgets['format'] = QComboBox()
        self.filter_widgets['format'].addItem("All Formats")
        self.filter_widgets['format'].addItem("ZIP")
        self.filter_widgets['format'].addItem("TAR.GZ")
        self.filter_widgets['format'].addItem("TAR")
        self.filter_widgets['format'].addItem("XLSX")
        self.filter_widgets['format'].setFixedWidth(120)
        self.filter_widgets['format'].currentTextChanged.connect(self.apply_filters)
        filter_layout.addWidget(QLabel("Format:"))
        filter_layout.addWidget(self.filter_widgets['format'])

        # Status filter (ComboBox)
        self.filter_widgets['status'] = QComboBox()
        self.filter_widgets['status'].addItem("All")
        self.filter_widgets['status'].addItem("Pass")
        self.filter_widgets['status'].addItem("Fail")
        self.filter_widgets['status'].setFixedWidth(100)
        self.filter_widgets['status'].currentTextChanged.connect(self.apply_filters)
        filter_layout.addWidget(QLabel("Status:"))
        filter_layout.addWidget(self.filter_widgets['status'])

        # Details filter
        self.filter_widgets['details'] = QLineEdit()
        self.filter_widgets['details'].setPlaceholderText("Details/Notes...")
        self.filter_widgets['details'].setFixedWidth(150)
        self.filter_widgets['details'].textChanged.connect(self.apply_filters)
        filter_layout.addWidget(QLabel("Details:"))
        filter_layout.addWidget(self.filter_widgets['details'])

        # Clear filters button
        self.clear_filter_btn = QPushButton("Clear Filters")
        self.clear_filter_btn.setStyleSheet("""
            QPushButton {
                background-color: #FF5722;
                color: white;
                border: none;
                padding: 4px 12px;
                border-radius: 3px;
                font-size: 10px;
            }
            QPushButton:hover {
                background-color: #E64A19;
            }
        """)
        self.clear_filter_btn.clicked.connect(self.clear_filters)
        filter_layout.addWidget(self.clear_filter_btn)

        filter_layout.addStretch()
        layout.addLayout(filter_layout)

        # Filter result label
        self.filter_result_label = QLabel("")
        self.filter_result_label.setStyleSheet("color: #666; font-size: 10px;")
        layout.addWidget(self.filter_result_label)

        # Table
        self.table = QTableWidget()
        self.table.setColumnCount(7)
        self.table.setHorizontalHeaderLabels(["Filename", "Path", "Size", "Format", "Package Type", "Status", "Details"])

        # Set column widths
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(0, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(1, QHeaderView.ResizeMode.Stretch)
        header.setSectionResizeMode(2, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(3, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(4, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(5, QHeaderView.ResizeMode.ResizeToContents)
        header.setSectionResizeMode(6, QHeaderView.ResizeMode.ResizeToContents)  # Details 列也改为 ResizeToContents

        # 设置 Details 列最小宽度，避免内容被挤压
        header.resizeSection(6, 300)

        # Set table style
        self.table.setAlternatingRowColors(True)
        self.table.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)
        self.table.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
        self.table.verticalHeader().setVisible(False)

        # 启用 word wrap，让长文本自动换行
        self.table.setWordWrap(True)

        # Set cursor style for clickable "View Details" links
        self.table.setStyleSheet("""
            QTableWidget {
                selection-background-color: #BBDEFB;
            }
            QTableWidget::item:hover {
                background-color: #E3F2FD;
            }
            QTableWidget::item:selected {
                background-color: #BBDEFB;
            }
        """)

        # Connect row double click event to show details (only for failed/warning packages)
        self.table.itemDoubleClicked.connect(self.on_table_double_clicked)

        layout.addWidget(self.table)

        # Bottom button area
        button_layout = QHBoxLayout()

        self.export_btn = QPushButton("Export Results")
        self.export_btn.setFixedWidth(SIZES['button_medium'])
        self.export_btn.setStyleSheet(f"""
            QPushButton {{
                background-color: {COLORS['warning']};
                color: {COLORS['text_inverse']};
                border: none;
                padding: {SPACING['small']}px {SPACING['medium']}px;
                border-radius: {BORDER_RADIUS['tiny']}px;
            }}
            QPushButton:hover {{
                background-color: #F57C00;
            }}
        """)
        self.export_btn.clicked.connect(self.export_results)
        self.export_btn.setEnabled(False)
        button_layout.addWidget(self.export_btn)

        button_layout.addStretch()

        self.stats_label = QLabel("Ready")
        button_layout.addWidget(self.stats_label)

        layout.addLayout(button_layout)

        # Status bar
        self.statusBar = QStatusBar()
        self.setStatusBar(self.statusBar)

    def select_directory(self):
        """Select directory"""
        directory = QFileDialog.getExistingDirectory(
            self,
            "Select Package Directory",
            "",
            QFileDialog.Option.ShowDirsOnly
        )

        if directory:
            self.current_dir = directory
            self.dir_path_label.setText(f"📁 {directory}")
            self.validate_btn.setEnabled(True)
            self.statusBar.showMessage(f"Selected directory: {directory}")

    def start_validation(self):
        """Start validation"""
        if not self.current_dir:
            QMessageBox.warning(self, "Warning", "Please select a data package directory first")
            return

        # Clear table
        self.table.setRowCount(0)
        self.stats_label.setText("Scanning...")
        self.validate_btn.setEnabled(False)
        self.select_btn.setEnabled(False)
        self.progress_bar.setVisible(True)
        self.progress_bar.setValue(0)

        # Start validation thread
        self.validation_thread = ValidationThread(self.current_dir)
        self.validation_thread.finished.connect(self.on_validation_finished)
        self.validation_thread.progress.connect(self.update_progress)
        self.validation_thread.start()

    def update_progress(self, value):
        """Update progress"""
        self.progress_bar.setValue(value)

    def on_validation_finished(self, scanner, error):
        """Validation finished"""
        self.validate_btn.setEnabled(True)
        self.select_btn.setEnabled(True)
        self.progress_bar.setVisible(False)

        if error:
            # 根据异常类型显示不同的错误提示
            if isinstance(error, ValidationException):
                title = "Error"
                message = f"{error.message}\n\n💡 Suggestion: {error.suggestion}"
                icon = QMessageBox.Icon.Warning
                self.stats_label.setText("Validation failed")

            elif isinstance(error, str):
                # 未知错误（字符串形式的错误信息）
                title = "Unknown Error"
                message = f"An unknown error occurred during validation:\n\n{error}\n\n" \
                         "Please contact the support team."
                icon = QMessageBox.Icon.Critical
                self.stats_label.setText("Validation failed")

            else:
                title = "Error"
                message = f"An error occurred during validation:\n\n{str(error)}"
                icon = QMessageBox.Icon.Critical
                self.stats_label.setText("Validation failed")

            QMessageBox.critical(self, title, message, icon)
            return

        self.scanner = scanner
        self.display_results()

    def display_results(self):
        """Display results"""
        stats = self.scanner.get_statistics()

        # Check if no files found
        if stats['total_files'] == 0:
            # Display empty state
            self.table.setRowCount(1)
            empty_item = QTableWidgetItem("No files found in the selected directory")
            empty_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            empty_item.setForeground(QColor(COLORS['text_hint']))
            empty_item.setFont(QFont("Arial", 12, QFont.Weight.Normal))
            self.table.setItem(0, 0, empty_item)
            self.table.setSpan(0, 0, 1, 7)  # Merge all columns

            # Update statistics
            self.stats_label.setText("Total: 0 files | Pass: 0 | Fail: 0")
            self.statusBar.showMessage("Scan completed: No files found")
            self.export_btn.setEnabled(False)

            QMessageBox.information(
                self,
                "Scan Completed",
                "No files found in the selected directory.\n\n"
                "Please select a different directory that contains data package files."
            )
            return

        # 保存原始数据用于筛选
        self.all_valid_files = []
        for file_info in self.scanner.valid_files:
            # Get package type and details
            package_type = file_info.get('package_type', 'Unknown')
            package_details = file_info.get('package_details', {})

            # Set color based on package type
            if package_type == 'NIC Package':
                package_color = COLORS['primary']
            elif package_type == 'Unknown':
                package_color = COLORS['warning']
            else:
                package_color = COLORS['text_secondary']

            # Generate detail information
            # Check if Package Type is Unknown
            if package_type == 'Unknown':
                status = 'Fail'
                status_color = COLORS['error']
                detail = 'Invalid package format'
            else:
                # Start with Pass status
                status = 'Pass'
                status_color = COLORS['success']
                detail = ''

                # Check if there are any issues (package-level or NE-level)
                has_issues = False

                if package_type == 'NIC Package':
                    nic_validation = file_info.get('nic_validation')
                    if nic_validation:
                        # Check package-level issues
                        if not nic_validation.get('valid', True):
                            has_issues = True
                            status = 'Fail'
                            status_color = COLORS['error']

                        # Check NE-level issues (change Warning to Fail)
                        static_mml_validation = nic_validation.get('static_mml_validation')
                        if static_mml_validation:
                            invalid_ne = static_mml_validation.get('invalid_ne_count', 0)
                            if invalid_ne > 0:
                                has_issues = True
                                status = 'Fail'  # Changed from Warning to Fail
                                status_color = COLORS['error']

                        # Check scenario validation issues
                        scenario_validation = nic_validation.get('scenario_validation')
                        if scenario_validation:
                            invalid_scenario = scenario_validation.get('invalid_ne_count', 0)
                            if invalid_scenario > 0:
                                has_issues = True
                                status = 'Fail'
                                status_color = COLORS['error']

                # Only show "View Details" button for failed packages
                if has_issues:
                    detail = 'View Details'

            self.all_valid_files.append({
                'file_info': file_info,
                'name': file_info['name'],
                'path': file_info['relative_path'],
                'size': self._format_size(file_info['size']),
                'format': self._get_file_type(file_info['name']),
                'package_type': package_type,
                'status': status,
                'detail': detail,
                'status_color': status_color,
                'package_color': package_color
            })

        # 保存无效文件数据
        self.all_invalid_files = []
        for file_info in self.scanner.illegal_files:
            self.all_invalid_files.append({
                'file_info': file_info,
                'name': file_info['name'],
                'path': file_info['relative_path'],
                'size': self._format_size(file_info['size']),
                'format': self._get_file_type(file_info['name']),
                'package_type': '-',
                'status': 'Fail',
                'detail': 'Invalid file format',
                'status_color': '#F44336',
                'package_color': '#666666'
            })

        # 应用筛选并显示
        self.apply_filters()

        # Update statistics
        self.stats_label.setText(
            f"Total: {stats['total_files']} files | "
            f"Pass: {stats['valid_files']} | "
            f"Fail: {stats['illegal_files']}"
        )

        self.statusBar.showMessage(f"Scan completed: {stats['total_files']} files found")
        self.export_btn.setEnabled(True)

        # Show alert
        if stats['illegal_files'] > 0:
            QMessageBox.warning(
                self,
                "Scan Completed",
                f"Scan completed!\n\n"
                f"Total files: {stats['total_files']}\n"
                f"Pass: {stats['valid_files']}\n"
                f"Fail: {stats['illegal_files']}\n\n"
                f"Please check the failed files in the table."
            )
        else:
            QMessageBox.information(
                self,
                "Scan Completed",
                f"Scan completed!\n\n"
                f"All {stats['total_files']} files have valid format."
            )

    def add_file_row(self, name, path, size, file_type, package_type, status, detail, status_color, package_color):
        """Add file row to table"""
        row = self.table.rowCount()
        self.table.insertRow(row)

        items = [name, path, size, file_type, package_type, status, detail]

        for col, item_text in enumerate(items):
            item = QTableWidgetItem(str(item_text))
            # Status column (column 5) uses status color
            if col == 5:
                item.setForeground(QColor(status_color))
                item.setFont(QFont("Arial", 9, QFont.Weight.Bold))
            # Package type column (column 4) uses package type color
            elif col == 4:
                item.setForeground(QColor(package_color))
                item.setFont(QFont("Arial", 9))
            # Details column (column 6): add tooltip, enable word wrap, and make clickable
            elif col == 6:
                item.setToolTip(str(item_text))  # Tooltip 显示完整内容
                # 启用 word wrap
                item.setTextAlignment(Qt.AlignmentFlag.AlignLeft | Qt.AlignmentFlag.AlignTop)

                # Only apply link style to "View Details" text
                if str(item_text) == 'View Details':
                    item.setForeground(QColor(COLORS['link']))
                    font = QFont("Arial", 9)
                    font.setUnderline(True)
                    item.setFont(font)

            self.table.setItem(row, col, item)

    def _format_size(self, size):
        """Format file size"""
        if size < 1024:
            return f"{size} B"
        elif size < 1024 * 1024:
            return f"{size/1024:.1f} KB"
        else:
            return f"{size/(1024*1024):.1f} MB"

    def _get_file_type(self, filename):
        """Get file type"""
        if filename.endswith('.zip'):
            return 'ZIP'
        elif filename.endswith('.tar.gz'):
            return 'TAR.GZ'
        elif filename.endswith('.tar'):
            return 'TAR'
        elif filename.endswith('.xlsx'):
            return 'XLSX'
        else:
            return 'Unknown'

    def apply_filters(self):
        """应用筛选条件并更新表格显示"""
        # 获取筛选条件
        filename_filter = self.filter_widgets['filename'].text().lower().strip()
        format_filter = self.filter_widgets['format'].currentText()
        status_filter = self.filter_widgets['status'].currentText()
        details_filter = self.filter_widgets['details'].text().lower().strip()

        # 清空表格
        self.table.setRowCount(0)

        # 筛选并显示有效文件
        for file_data in self.all_valid_files:
            if self._matches_filter(file_data, filename_filter, format_filter, status_filter, details_filter):
                self.add_file_row(
                    file_data['name'],
                    file_data['path'],
                    file_data['size'],
                    file_data['format'],
                    file_data['package_type'],
                    file_data['status'],
                    file_data['detail'],
                    file_data['status_color'],
                    file_data['package_color']
                )

        # 筛选并显示无效文件
        for file_data in self.all_invalid_files:
            if self._matches_filter(file_data, filename_filter, format_filter, status_filter, details_filter):
                self.add_file_row(
                    file_data['name'],
                    file_data['path'],
                    file_data['size'],
                    file_data['format'],
                    file_data['package_type'],
                    file_data['status'],
                    file_data['detail'],
                    file_data['status_color'],
                    file_data['package_color']
                )

        # 更新筛选结果标签
        total_files = len(self.all_valid_files) + len(self.all_invalid_files)
        displayed_files = self.table.rowCount()
        if displayed_files == total_files:
            self.filter_result_label.setText(f"Showing all {total_files} records")
        else:
            self.filter_result_label.setText(f"Showing {displayed_files} of {total_files} records")

    def _matches_filter(self, file_data, filename_filter, format_filter, status_filter, details_filter):
        """检查文件数据是否匹配筛选条件"""
        # Filename filter (文本模糊匹配)
        if filename_filter and filename_filter not in file_data['name'].lower():
            return False

        # Format filter
        if format_filter != "All Formats" and file_data['format'] != format_filter:
            return False

        # Status filter
        if status_filter != "All" and file_data['status'] != status_filter:
            return False

        # Details filter (文本模糊匹配)
        if details_filter and details_filter not in file_data['detail'].lower():
            return False

        return True

    def clear_filters(self):
        """清除所有筛选条件"""
        self.filter_widgets['filename'].clear()
        self.filter_widgets['format'].setCurrentIndex(0)  # All Formats
        self.filter_widgets['status'].setCurrentIndex(0)  # All
        self.filter_widgets['details'].clear()
        self.apply_filters()

    def export_results(self):
        """Export results"""
        if not self.scanner:
            QMessageBox.warning(self, "Warning", "No results to export")
            return

        # Select save file
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Results",
            f"validation_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel files (*.xlsx);;Text files (*.txt);;CSV files (*.csv);;All files (*.*)"
        )

        if file_path:
            try:
                if file_path.endswith('.xlsx'):
                    self._export_excel(file_path)
                elif file_path.endswith('.csv'):
                    self._export_csv(file_path)
                else:
                    self._export_txt(file_path)

                QMessageBox.information(self, "Success", f"Results exported to:\n{file_path}")
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Export failed:\n{str(e)}")

    def _export_txt(self, filename):
        """Export to text file"""
        with open(filename, 'w', encoding='utf-8') as f:
            f.write("Network Package Validation Results\n")
            f.write("=" * 50 + "\n")
            f.write(f"Validation time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"Package directory: {self.current_dir}\n")
            f.write(f"Supported formats: ZIP, TAR.GZ, TAR, XLSX\n")
            f.write("\n")

            # Statistics
            stats = self.scanner.get_statistics()
            if stats:
                f.write(f"Statistics:\n")
                f.write(f"  Total files: {stats['total_files']}\n")
                f.write(f"  Pass: {stats['valid_files']}\n")
                f.write(f"  Fail: {stats['illegal_files']}\n")
                f.write("\n")

            # Detailed results
            f.write("Detailed Results:\n")
            f.write("-" * 50 + "\n")

            for row in range(self.table.rowCount()):
                f.write(f"Filename: {self.table.item(row, 0).text()}\n")
                f.write(f"Path: {self.table.item(row, 1).text()}\n")
                f.write(f"Size: {self.table.item(row, 2).text()}\n")
                f.write(f"Format: {self.table.item(row, 3).text()}\n")
                f.write(f"Package Type: {self.table.item(row, 4).text()}\n")
                f.write(f"Status: {self.table.item(row, 5).text()}\n")
                detail = self.table.item(row, 6).text()
                if detail:
                    f.write(f"Details: {detail}\n")
                f.write("\n")

    def on_table_double_clicked(self, item):
        """Handle table double click to show details dialog (only for failed/warning packages)"""
        # Get row
        row = self.table.row(item)
        file_name = self.table.item(row, 0).text()

        # Find the corresponding file data
        file_data = None
        for data in self.all_valid_files + self.all_invalid_files:
            if data['name'] == file_name:
                file_data = data
                break

        if not file_data:
            return

        # Check status - only show details for failed/warning packages
        status = file_data.get('status', '')
        if status == 'Pass':
            # Pass package - no details needed
            return

        # Show details dialog for failed/warning packages
        self.show_details_dialog(file_data)

    def show_details_dialog(self, file_data):
        """Show details dialog for failed/warning packages"""
        # Create dialog
        from PyQt6.QtWidgets import QDialog

        dialog = QDialog(self)
        dialog.setWindowTitle(f"Validation Details - {file_data.get('name', '')}")
        dialog.setMinimumSize(950, 700)
        dialog.setStyleSheet("""
            QDialog {
                background-color: #f5f5f5;
            }
        """)

        layout = QVBoxLayout()
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(16)

        # Package basic info
        file_info = file_data.get('file_info', {})
        package_type = file_data.get('package_type', 'Unknown')

        info_html = f"""
        <div style="background-color: #E3F2FD; padding: 12px; border-radius: 6px;">
            <h3 style="margin: 0 0 12px 0; color: #1976D2;">{file_data.get('name', '')}</h3>
            <table style="border-collapse: collapse;">
                <tr>
                    <td style="padding: 6px 12px 6px 0; color: #555;"><strong>Path:</strong> {file_data.get('path', '')}</td>
                </tr>
                <tr>
                    <td style="padding: 6px 12px 6px 0; color: #555;"><strong>Size:</strong> {file_data.get('size', '')}</td>
                </tr>
                <tr>
                    <td style="padding: 6px 12px 12px 0; color: #555;"><strong>Package Type:</strong> <span style="color: #2196F3; font-weight: bold;">{package_type}</span></td>
                </tr>
            </table>
        </div>
        """

        info_label = QLabel()
        info_label.setTextFormat(Qt.TextFormat.RichText)
        info_label.setWordWrap(True)
        info_label.setText(info_html)
        layout.addWidget(info_label)

        # Add package-level and NE-level issues for NIC packages
        if package_type == 'NIC Package':
            nic_validation = file_info.get('nic_validation')
            if nic_validation:
                # Package-level issues table
                package_issues_html = self._get_package_issues_table_html(nic_validation)
                layout.addWidget(package_issues_html)

                # NE-level issues table
                ne_issues_html = self._get_ne_issues_table_html(nic_validation)
                layout.addWidget(ne_issues_html)

        # Close button
        button_layout = QVBoxLayout()
        button_layout.setContentsMargins(0, 20, 0, 0)

        close_btn = QPushButton("Close")
        close_btn.setFixedWidth(120)
        close_btn.setStyleSheet("""
            QPushButton {
                background-color: #2196F3;
                color: white;
                border: none;
                padding: 10px 24px;
                border-radius: 6px;
                font-weight: bold;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #1976D2;
            }
            QPushButton:pressed {
                background-color: #1565C0;
            }
        """)
        close_btn.clicked.connect(dialog.accept)
        button_layout.addWidget(close_btn)
        button_layout.addStretch()
        layout.addLayout(button_layout)

        dialog.setLayout(layout)
        dialog.exec()

    def _get_package_issues_table_html(self, nic_validation):
        """Generate QLabel with package-level issues table"""
        html = f"""
        <div style="border: 2px solid #FFCDD2; border-radius: 6px; overflow: hidden;">
            <div style="background-color: #FFEBEE; padding: 12px; font-weight: bold; color: #C62828;">
                Package Level Issues
            </div>
            <table style="width: 100%; border-collapse: collapse; margin: 12px;">
                <tr style="background-color: #f9f9f9; border-bottom: 2px solid #FFCDD2;">
                    <th style="padding: 10px; text-align: left; color: #555;">Type</th>
                    <th style="padding: 10px; text-align: left; color: #555;">Description</th>
                </tr>
        """

        issues = []

        # Check collect range
        if nic_validation.get('collect_range_too_short'):
            hours = nic_validation.get('collect_range_hours', 0)
            issues.append(('Warning', f'Collection time range is too short ({hours:.2f}h < 24h)'))

        # Check anonymous mode
        if nic_validation.get('anonymous_mode_invalid'):
            issues.append(('Warning', 'Collected in anonymous mode'))

        # Check missing neinfo
        if nic_validation.get('missing_neinfo'):
            issues.append(('Error', 'Missing required file neinfo.txt'))

        # Check missing folders
        missing_folders = nic_validation.get('missing_folders', [])
        if missing_folders:
            issues.append(('Warning', f'Missing {len(missing_folders)} NE data folders'))

        # Skip NE-level issues (missing files, unsupported types) - these are shown in NE details table

        # Add rows
        for idx, (issue_type, description) in enumerate(issues, 1):
            row_style = ''
            if idx % 2 == 0:
                row_style = 'background-color: #ffffff;'
            else:
                row_style = 'background-color: #fafafa;'

            color = '#F44336' if issue_type == 'Error' else '#FF9800'

            html += f"""
                <tr style="{row_style} border-bottom: 1px solid #e0e0e0;">
                    <td style="padding: 10px; font-weight: bold; color: {color};">{issue_type}</td>
                    <td style="padding: 10px; color: #555;">{description}</td>
                </tr>
            """

        if not issues:
            html += f"""
                <tr style="background-color: #ffffff;">
                    <td colspan="2" style="padding: 12px; color: #4CAF50; font-weight: bold; text-align: center;">
                        ✓ No package-level issues found
                    </td>
                </tr>
            """

        html += """
            </table>
        </div>
        """

        label = QLabel()
        label.setTextFormat(Qt.TextFormat.RichText)
        label.setWordWrap(True)
        label.setText(html)
        return label

    def _get_ne_issues_table_html(self, nic_validation):
        """Generate QLabel with NE-level issues table"""
        html = """
        <div style="border: 2px solid #E3F2FD; border-radius: 6px; overflow: hidden;">
            <div style="background-color: #E3F2FD; padding: 12px; font-weight: bold; color: #1976D2;">
                Network Element Details
            </div>
            <table style="width: 100%; border-collapse: collapse; margin: 12px;">
                <tr style="background-color: #f9f9f9; border-bottom: 2px solid #E3F2FD;">
                    <th style="padding: 10px; text-align: left; color: #555;">#</th>
                    <th style="padding: 10px; text-align: left; color: #555;">NE Name (Type)</th>
                    <th style="padding: 10px; text-align: left; color: #555;">Status</th>
                    <th style="padding: 10px; text-align: left; color: #555;">Error Type</th>
                    <th style="padding: 10px; text-align: left; color: #555;">Issues</th>
                </tr>
        """

        static_mml_validation = nic_validation.get('static_mml_validation')
        scenario_validation = nic_validation.get('scenario_validation')

        # Build a dictionary to combine results from both validators
        ne_combined_results = {}

        # Process static MML results
        if static_mml_validation:
            ne_results = static_mml_validation.get('ne_results', [])
            for ne_result in ne_results:
                ne_key = (ne_result.get('ne_name'), ne_result.get('ne_type'))
                ne_combined_results[ne_key] = {
                    'static_mml': ne_result,
                    'scenario': None
                }

        # Process scenario results
        if scenario_validation:
            ne_results = scenario_validation.get('ne_results', [])
            for ne_result in ne_results:
                ne_key = (ne_result.get('ne_name'), ne_result.get('ne_type'))
                if ne_key in ne_combined_results:
                    ne_combined_results[ne_key]['scenario'] = ne_result
                else:
                    ne_combined_results[ne_key] = {
                        'static_mml': None,
                        'scenario': ne_result
                    }

        # Generate table rows
        if ne_combined_results:
            row_idx = 0
            for ne_key, results in ne_combined_results.items():
                ne_name, ne_type = ne_key
                static_mml_result = results.get('static_mml')
                scenario_result = results.get('scenario')

                # Collect all errors for this NE
                errors = []

                # Check static MML result
                if static_mml_result:
                    static_valid = static_mml_result.get('valid')
                    if static_valid is False:
                        missing_paths = static_mml_result.get('missing_paths', [])
                        if missing_paths:
                            for path in missing_paths:
                                errors.append(('Static MML Missing', f'<span style="color: #F44336;">{path}</span>'))

                # Check scenario result
                if scenario_result:
                    scenario_valid = scenario_result.get('valid')
                    if scenario_valid is False:
                        scenario_error_msg = scenario_result.get('error')
                        if scenario_error_msg:
                            errors.append(('Scenario Error', f'<span style="color: #FF9800;">{scenario_error_msg}</span>'))

                # Determine overall status
                status = 'Pass'
                status_color = '#4CAF50'
                if errors:
                    status = 'Fail'
                    status_color = '#F44336'

                # Calculate row count for rowspan
                num_rows = max(1, len(errors))

                # For each error type (or one row if no errors)
                for error_row_num in range(num_rows):
                    row_style = 'background-color: #ffffff;' if row_idx % 2 == 1 else 'background-color: #fafafa;'

                    # First 3 columns with rowspan
                    if error_row_num == 0:
                        first_col = f'<td style="padding: 10px; color: #666;" rowspan="{num_rows}">{row_idx + 1}</td>'
                        second_col = f'<td style="padding: 10px; color: #555; font-weight: bold;" rowspan="{num_rows}">{ne_name}<br><span style="font-weight: normal; font-size: 0.85em;">({ne_type})</span></td>'
                        third_col = f'<td style="padding: 10px; font-weight: bold; color: {status_color};" rowspan="{num_rows}">{status}</td>'
                    else:
                        first_col = ''
                        second_col = ''
                        third_col = ''

                    # Error Type and Issues columns
                    if errors:
                        error_type, error_msg = errors[error_row_num]
                        fourth_col = f'<td style="padding: 10px; color: #555;">{error_type}</td>'
                        fifth_col = f'<td style="padding: 10px; color: #555;">{error_msg}</td>'
                    else:
                        fourth_col = '<td style="padding: 10px; color: #555;">-</td>'
                        fifth_col = '<td style="padding: 10px; color: #555;">-</td>'

                    html += f"""
                        <tr style="{row_style} border-bottom: 1px solid #e0e0e0;">
                            {first_col}
                            {second_col}
                            {third_col}
                            {fourth_col}
                            {fifth_col}
                        </tr>
                    """
                    row_idx += 1

            # Add summary row
            total_ne = static_mml_validation.get('total_ne_count', 0) if static_mml_validation else 0
            valid_ne = static_mml_validation.get('valid_ne_count', 0) if static_mml_validation else 0
            invalid_ne = static_mml_validation.get('invalid_ne_count', 0) if static_mml_validation else 0
            warning_ne = static_mml_validation.get('warning_ne_count', 0) if static_mml_validation else 0

            # Add scenario validation count
            scenario_invalid = scenario_validation.get('invalid_ne_count', 0) if scenario_validation else 0

            html += f"""
                <tr style="background-color: #F5F5F5; border-top: 2px solid #E3F2FD;">
                    <td colspan="5" style="padding: 12px; font-weight: bold;">
                        Summary: Total: {total_ne} |
                        <span style="color: #4CAF50;">Pass: {valid_ne}</span> |
                        <span style="color: #FF9800;">Warning: {warning_ne}</span> |
                        <span style="color: #F44336;">Static MML Fail: {invalid_ne}</span> |
                        <span style="color: #FF9800;">Scenario Fail: {scenario_invalid}</span>
                    </td>
                </tr>
            """
        else:
            html += f"""
                <tr style="background-color: #ffffff;">
                    <td colspan="5" style="padding: 12px; color: #666; text-align: center;">
                        No NE validation data available
                    </td>
                    </tr>
                """

        html += """
            </table>
        </div>
        """

        label = QLabel()
        label.setTextFormat(Qt.TextFormat.RichText)
        label.setWordWrap(True)
        label.setText(html)
        return label

    def _export_csv(self, filename):
        """Export to CSV file"""
        import csv

        with open(filename, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['Filename', 'Path', 'Size', 'Format', 'Package Type', 'Status', 'Details'])

            for row in range(self.table.rowCount()):
                writer.writerow([
                    self.table.item(row, 0).text(),
                    self.table.item(row, 1).text(),
                    self.table.item(row, 2).text(),
                    self.table.item(row, 3).text(),
                    self.table.item(row, 4).text(),
                    self.table.item(row, 5).text(),
                    self.table.item(row, 6).text()
                ])

    def _export_excel(self, filename):
        """Export to Excel file with multiple worksheets"""
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl is not installed. Install it with: pip install openpyxl")

        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        # Define styles
        header_font = Font(bold=True, color='333333')
        header_fill = PatternFill(start_color='E3F2FD', end_color='E3F2FD', fill_type='solid')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # 1. Summary Sheet
        ws_summary = wb.create_sheet('Summary')
        self._create_summary_sheet(ws_summary, header_font, header_fill, border, center_alignment)

        # 2. Package List Sheet
        ws_package = wb.create_sheet('Package List')
        self._create_package_list_sheet(ws_package, header_font, header_fill, border, left_alignment)

        # 3. NE Validation Sheet
        ws_ne_validation = wb.create_sheet('NE Validation')
        self._create_ne_validation_sheet(ws_ne_validation, header_font, header_fill, border, left_alignment)

        # 4. NE Errors Sheet
        ws_ne_errors = wb.create_sheet('NE Errors')
        self._create_ne_errors_sheet(ws_ne_errors, header_font, header_fill, border, left_alignment)

        # Save workbook
        wb.save(filename)

    def _create_summary_sheet(self, ws, header_font, header_fill, border, center_alignment):
        """Create Summary sheet with statistics"""
        # Headers
        headers = ['统计项', '数值']
        ws.append(headers)

        for col in range(1, len(headers) + 1):
            cell = ws.cell(1, col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = border

        # Data
        all_files = self.all_valid_files + self.all_invalid_files
        total_files = len(all_files)
        pass_count = len([f for f in all_files if f.get('status') == 'Pass'])
        fail_count = len([f for f in all_files if f.get('status') == 'Fail'])

        row = 2
        ws.append(['验证时间', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        ws.append(['扫描目录', self.current_dir or '-'])
        ws.append([])
        ws.append(['包总数', total_files])
        ws.append(['Pass', pass_count])
        ws.append(['Fail', fail_count])
        ws.append([])
        row += 8

        # NE statistics - accumulate from all files
        total_ne = 0
        ne_pass = 0
        ne_static_mml_fail = 0
        ne_scenario_fail = 0

        for file_data in self.all_valid_files + self.all_invalid_files:
            file_info = file_data.get('file_info', {})
            nic_validation = file_info.get('nic_validation', {})
            static_mml_validation = nic_validation.get('static_mml_validation')
            scenario_validation = nic_validation.get('scenario_validation')

            if static_mml_validation:
                total_ne += static_mml_validation.get('total_ne_count', 0)
                ne_pass += static_mml_validation.get('valid_ne_count', 0)
                ne_static_mml_fail += static_mml_validation.get('invalid_ne_count', 0)
                ne_scenario_fail += scenario_validation.get('invalid_ne_count', 0) if scenario_validation else 0

        ws.append(['网元总数', total_ne])
        ws.append(['网元 Pass', ne_pass])
        ws.append(['网元 Static MML Fail', ne_static_mml_fail])
        ws.append(['网元 Scenario Fail', ne_scenario_fail])

        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30

    def _create_package_list_sheet(self, ws, header_font, header_fill, border, left_alignment):
        """Create Package List sheet with all package information"""
        headers = ['文件名', '路径', '大小', '格式', '包类型', '状态', '异常信息']
        ws.append(headers)

        for col in range(1, len(headers) + 1):
            cell = ws.cell(1, col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = left_alignment
            cell.border = border

        # Data rows
        row = 2
        for file_data in self.all_valid_files + self.all_invalid_files:
            status = file_data.get('status', 'Unknown')
            # Extract error message from validation result
            file_info = file_data.get('file_info', {})
            nic_validation = file_info.get('nic_validation', {})
            error_msg = '-'

            # Check for package-level errors
            if nic_validation:
                errors = nic_validation.get('errors', [])
                if errors and not nic_validation.get('valid', True):
                    error_msg = errors[0] if isinstance(errors[0], str) else str(errors[0])
                else:
                    # Check for NE-level errors
                    static_mml_validation = nic_validation.get('static_mml_validation')
                    scenario_validation = nic_validation.get('scenario_validation')
                    invalid_ne_count = 0
                    if static_mml_validation:
                        invalid_ne_count += static_mml_validation.get('invalid_ne_count', 0)
                    if scenario_validation:
                        invalid_ne_count += scenario_validation.get('invalid_ne_count', 0)
                    if invalid_ne_count > 0:
                        error_msg = f'{invalid_ne_count} NE(s) have validation errors'

            ws.append([
                file_data.get('name', ''),
                file_data.get('path', ''),
                file_data.get('size', ''),
                file_data.get('format', ''),
                file_data.get('package_type', ''),
                status,
                error_msg
            ])

            # Set status color
            status_cell = ws.cell(row, 6)
            if status == 'Pass':
                status_cell.font = Font(color='008000', bold=True)
            elif status == 'Fail':
                status_cell.font = Font(color='CC0000', bold=True)
            elif status == 'Warning':
                status_cell.font = Font(color='FF9800', bold=True)

            # Apply border to all cells
            for col in range(1, 8):
                ws.cell(row, col).border = border
                ws.cell(row, col).alignment = left_alignment

            row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 8
        ws.column_dimensions['G'].width = 50

    def _create_ne_validation_sheet(self, ws, header_font, header_fill, border, left_alignment):
        """Create NE Validation sheet with merged cells"""
        headers = ['包文件名', '#', 'NE Name', 'NE Type', 'Status', 'Error Type', 'Issues']
        ws.append(headers)

        for col in range(1, len(headers) + 1):
            cell = ws.cell(1, col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = left_alignment
            cell.border = border

        row = 2
        idx = 1

        for file_data in self.all_valid_files + self.all_invalid_files:
            file_name = file_data.get('name', '')
            file_info = file_data.get('file_info', {})
            nic_validation = file_info.get('nic_validation', {})
            static_mml_validation = nic_validation.get('static_mml_validation')
            scenario_validation = nic_validation.get('scenario_validation')

            if not static_mml_validation and not scenario_validation:
                continue

            # Combine results from both validators
            ne_combined_results = {}

            if static_mml_validation:
                for ne_result in static_mml_validation.get('ne_results', []):
                    ne_key = (ne_result.get('ne_name'), ne_result.get('ne_type'))
                    ne_combined_results[ne_key] = {
                        'static_mml': ne_result,
                        'scenario': None
                    }

            if scenario_validation:
                for ne_result in scenario_validation.get('ne_results', []):
                    ne_key = (ne_result.get('ne_name'), ne_result.get('ne_type'))
                    if ne_key in ne_combined_results:
                        ne_combined_results[ne_key]['scenario'] = ne_result
                    else:
                        ne_combined_results[ne_key] = {
                            'static_mml': None,
                            'scenario': ne_result
                        }

            # Generate rows
            for ne_key, results in ne_combined_results.items():
                ne_name, ne_type = ne_key
                static_mml_result = results.get('static_mml')
                scenario_result = results.get('scenario')

                # Collect errors
                errors = []

                if static_mml_result and not static_mml_result.get('valid'):
                    missing_paths = static_mml_result.get('missing_paths', [])
                    for path in missing_paths:
                        errors.append(('Static MML Missing', path))

                if scenario_result and not scenario_result.get('valid'):
                    error_msg = scenario_result.get('error', '')
                    if error_msg:
                        errors.append(('Scenario Error', error_msg))

                # Determine status
                status = 'Pass'
                status_color = '008000'
                if errors:
                    status = 'Fail'
                    status_color = 'CC0000'

                num_rows = max(1, len(errors))

                # First 4 columns with rowspan (merged cells) including package name
                for error_row_num in range(num_rows):
                    if error_row_num == 0:
                        # Add cells that will be merged
                        ws.cell(row, 1, file_name)
                        ws.cell(row, 2, idx)
                        ws.cell(row, 3, ne_name)
                        ws.cell(row, 4, ne_type)
                        ws.cell(row, 5, status)

                        # Merge cells for rowspan
                        if num_rows > 1:
                            ws.merge_cells(f'A{row}:A{row + num_rows - 1}')
                            ws.merge_cells(f'B{row}:B{row + num_rows - 1}')
                            ws.merge_cells(f'C{row}:C{row + num_rows - 1}')
                            ws.merge_cells(f'D{row}:D{row + num_rows - 1}')
                            ws.merge_cells(f'E{row}:E{row + num_rows - 1}')

                    # Error Type and Issues columns
                    if errors:
                        error_type, error_msg = errors[error_row_num]
                        ws.cell(row, 6, error_type)
                        ws.cell(row, 7, error_msg)
                    else:
                        ws.cell(row, 6, '-')
                        ws.cell(row, 7, '-')

                    # Apply styles
                    for col in range(1, 8):
                        cell = ws.cell(row, col)
                        cell.border = border
                        cell.alignment = left_alignment

                    # Status color (only first row, others are merged)
                    if error_row_num == 0:
                        ws.cell(row, 5).font = Font(color=status_color, bold=True)

                    row += 1

                idx += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 20  # 包文件名
        ws.column_dimensions['B'].width = 5   # #
        ws.column_dimensions['C'].width = 15  # NE Name
        ws.column_dimensions['D'].width = 12  # NE Type
        ws.column_dimensions['E'].width = 10  # Status
        ws.column_dimensions['F'].width = 20  # Error Type
        ws.column_dimensions['G'].width = 60  # Issues

    def _create_ne_errors_sheet(self, ws, header_font, header_fill, border, left_alignment):
        """Create NE Errors sheet with only errors"""
        headers = ['包文件名', 'NE Name', 'NE Type', 'Error Type', 'Issues']
        ws.append(headers)

        for col in range(1, len(headers) + 1):
            cell = ws.cell(1, col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = left_alignment
            cell.border = border

        row = 2

        for file_data in self.all_valid_files + self.all_invalid_files:
            file_name = file_data.get('name', '')
            file_info = file_data.get('file_info', {})
            nic_validation = file_info.get('nic_validation', {})
            static_mml_validation = nic_validation.get('static_mml_validation')
            scenario_validation = nic_validation.get('scenario_validation')

            if not static_mml_validation and not scenario_validation:
                continue

            # Collect NE errors
            ne_combined_results = {}

            if static_mml_validation:
                for ne_result in static_mml_validation.get('ne_results', []):
                    ne_key = (ne_result.get('ne_name'), ne_result.get('ne_type'))
                    if ne_key not in ne_combined_results:
                        ne_combined_results[ne_key] = {'static_mml': ne_result, 'scenario': None}
                    else:
                        ne_combined_results[ne_key]['static_mml'] = ne_result

            if scenario_validation:
                for ne_result in scenario_validation.get('ne_results', []):
                    ne_key = (ne_result.get('ne_name'), ne_result.get('ne_type'))
                    if ne_key not in ne_combined_results:
                        ne_combined_results[ne_key] = {'static_mml': None, 'scenario': ne_result}
                    else:
                        ne_combined_results[ne_key]['scenario'] = ne_result

            # Add error rows
            for ne_key, results in ne_combined_results.items():
                ne_name, ne_type = ne_key
                has_errors = False

                # Check static MML errors
                if results.get('static_mml') and not results['static_mml'].get('valid'):
                    missing_paths = results['static_mml'].get('missing_paths', [])
                    for path in missing_paths:
                        ws.append([file_name, ne_name, ne_type, 'Static MML Missing', path])
                        for col in range(1, 6):
                            ws.cell(row, col).border = border
                            ws.cell(row, col).alignment = left_alignment
                        row += 1
                    has_errors = True

                # Check scenario errors
                if results.get('scenario') and not results['scenario'].get('valid'):
                    error_msg = results['scenario'].get('error', '')
                    if error_msg:
                        ws.append([file_name, ne_name, ne_type, 'Scenario Error', error_msg])
                        for col in range(1, 6):
                            ws.cell(row, col).border = border
                            ws.cell(row, col).alignment = left_alignment
                        row += 1
                    has_errors = True

        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 60


def main():
    app = QApplication(sys.argv)
    window = ValidatorApp()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
